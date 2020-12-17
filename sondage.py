from itertools import islice

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

pd.options.mode.chained_assignment = None  # default='warn'

st.sidebar.image("logo/Logo_CITC_Gris.png", use_column_width=True)

wb = load_workbook(filename="Sondage.xlsx")
ws = wb.active
data = ws.values
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
# df = pd.read_excel("Sondage.xlsx", engine="openpyxl")

df = pd.DataFrame(data, index=idx, columns=cols)
df = df.iloc[1:]


wb2 = load_workbook(filename="Sondage_NL.xlsx")
ws2 = wb2.active
data2 = ws2.values
cols2 = next(data2)[1:]
data2 = list(data2)
idx2 = [r[0] for r in data2]
data2 = (islice(r, 1, None) for r in data2)
# df = pd.read_excel("Sondage.xlsx", engine="openpyxl")

df_nl = pd.DataFrame(data2, index=idx2, columns=cols2)

# df_nl = pd.read_excel("Sondage_NL.xlsx", engine="openpyxl")

st.title("Résumé du sondage factory 4.0")

cols1 = [
    c
    for c in df.columns
    if (c.lower()[:6] != "points" and c.lower()[:8] != "feedback")
]

df_redux = df[cols1[6:]]

cols2 = [
    c
    for c in df_nl.columns
    if (c.lower()[:6] != "points" and c.lower()[:8] != "feedback")
]
df_nl_redux = df_nl[cols2[6:]]

fr = st.sidebar.checkbox("Sondage francophone", value=True)

nl = st.sidebar.checkbox("Sondage flamand", value=False)

ne_se_prononce_pas = [
    "Si oui, lequel ?",
    "De quelle région votre entreprise provient-elle ?",
    "Pouvez-vous détailler ?",
    "Si vous avez déjà investi, quelle est la nature de cet investissement ?",
    "Dans le cas où vous avez choisi de ne pas investir, pouvez-vous nous préciser la raison ?",
    "Pouvez-vous détailler ?2",
    "Le programme d’accompagnement Factory 4.0 et/ou l’investissement associé ont-ils contribué faire monter en compétence certains employés ?",
    "Si oui, pouvez-vous préciser ?",
    "Le programme d’accompagnement Factory 4.0 et/ou l’investissement associé ont-ils eu un impact environnemental dans votre entreprise ?",
]
zero_nan = ["Combien ?", "Combien ?2"]


def fill_df(df):
    for c in df.columns:
        if c in ne_se_prononce_pas:
            df[c] = df[c].replace(np.nan, "Ne se prononce pas", regex=True)
        elif c in zero_nan:
            df[c] = df[c].fillna(0)
    return df


df_redux = fill_df(df_redux)


if fr:
    # st.write(list(df_redux.columns))

    # st.write(
    #     df_redux[
    #         "Comment qualifieriez-vous votre niveau de connaissances dans l’industrie 4.0, avant d’être accompagné par le programme Factory 4.0 ?"
    #     ].unique()
    # )

    st.dataframe(df_redux)

    st.header(
        "Sélectionnez une ou plusieurs questions pour obtenir des statistiques croisées"
    )
    questions = st.multiselect(
        "question", list(df_redux.columns), default=["Raison sociale"]
    )

    if questions == []:
        st.write("Vous devez au moins choisir une colonne !")
    else:
        vals = df_redux[questions].value_counts()
        st.table(vals)

if nl:
    st.dataframe(df_nl_redux)

    st.header(
        "Sélectionnez une ou plusieurs questions pour obtenir des statistiques croisées"
    )
    questions = st.multiselect(
        "question", list(df_nl_redux.columns), default=["Handelsnaam"]
    )

    if questions == []:
        st.write("Vous devez au moins choisir une colonne !")
    else:
        vals = df_nl_redux[questions].value_counts()
        st.table(vals)
